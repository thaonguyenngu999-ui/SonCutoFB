"""
Login Page - Đăng nhập Facebook cho các profiles
PySide6 version - WITH FULL AUTOMATION LOGIC
"""
import threading
import os
import queue
import time
import random
from typing import List, Dict, Optional
from datetime import datetime
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QFileDialog, QMessageBox, QTableWidgetItem, QSpinBox,
    QProgressBar
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QColor

from config import COLORS
from widgets import (
    CyberButton, CyberInput, CyberComboBox, CyberCard,
    CyberTitle, CyberStatCard, CyberTable, CyberCheckBox
)
from api_service import api
from db import get_profiles, sync_profiles
from automation.window_manager import (
    acquire_window_slot, release_window_slot, get_window_bounds
)

# Optional openpyxl
try:
    import openpyxl
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None

# Optional websocket
try:
    import websocket
    import requests
    HAS_WEBSOCKET = True
except ImportError:
    HAS_WEBSOCKET = False


class LoginPage(QWidget):
    """Login Page - Đăng nhập Facebook - FULL AUTOMATION"""

    def __init__(self, log_func, parent=None):
        super().__init__(parent)
        self.log = log_func
        self.folders: List[Dict] = []
        self.profiles: List[Dict] = []
        self.accounts: List[Dict] = []
        self.xlsx_path = ""
        self.workbook = None

        # Profile status: {uuid: {'has_fb': bool, 'status': str}}
        self.profile_status: Dict[str, Dict] = {}
        self.profile_checkboxes: Dict[str, CyberCheckBox] = {}

        # Running state
        self._is_running = False
        self._stop_requested = False

        self._setup_ui()
        QTimer.singleShot(500, self._load_folders)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)

        # ========== TOP BAR ==========
        top_bar = QHBoxLayout()
        top_bar.setSpacing(12)

        title = CyberTitle("Đăng nhập FB", "Đăng nhập Facebook tự động", "mint")
        top_bar.addWidget(title)

        top_bar.addStretch()

        self.stat_profiles = CyberStatCard("PROFILES", "0", "📁", "mint")
        self.stat_profiles.setFixedWidth(150)
        top_bar.addWidget(self.stat_profiles)

        self.stat_selected = CyberStatCard("ĐÃ CHỌN", "0", "✓", "cyan")
        self.stat_selected.setFixedWidth(150)
        top_bar.addWidget(self.stat_selected)

        self.stat_accounts = CyberStatCard("TÀI KHOẢN", "0", "👤", "purple")
        self.stat_accounts.setFixedWidth(150)
        top_bar.addWidget(self.stat_accounts)

        self.stat_live = CyberStatCard("LIVE", "0", "✅", "mint")
        self.stat_live.setFixedWidth(120)
        top_bar.addWidget(self.stat_live)

        self.stat_die = CyberStatCard("DIE", "0", "❌", "coral")
        self.stat_die.setFixedWidth(120)
        top_bar.addWidget(self.stat_die)

        layout.addLayout(top_bar)

        # ========== TOOLBAR ==========
        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)

        # Folder selection
        self.folder_combo = CyberComboBox(["📁 Chọn thư mục"])
        self.folder_combo.setFixedWidth(200)
        self.folder_combo.currentIndexChanged.connect(self._on_folder_change)
        toolbar.addWidget(self.folder_combo)

        btn_load = CyberButton("TẢI PROFILES", "cyan", "📥")
        btn_load.clicked.connect(self._load_profiles)
        toolbar.addWidget(btn_load)

        toolbar.addStretch()

        # Import XLSX
        self.xlsx_input = CyberInput("📂 Chọn file XLSX tài khoản...")
        self.xlsx_input.setFixedWidth(250)
        self.xlsx_input.setReadOnly(True)
        toolbar.addWidget(self.xlsx_input)

        btn_browse = CyberButton("CHỌN FILE", "purple", "📂")
        btn_browse.clicked.connect(self._browse_xlsx)
        toolbar.addWidget(btn_browse)

        toolbar.addStretch()

        btn_refresh = CyberButton("", "ghost", "🔄")
        btn_refresh.setFixedWidth(40)
        btn_refresh.clicked.connect(self._load_folders)
        toolbar.addWidget(btn_refresh)

        layout.addLayout(toolbar)

        # ========== MAIN CONTENT ==========
        content = QHBoxLayout()
        content.setSpacing(12)

        # LEFT - Profiles Table (rộng hơn)
        profiles_card = CyberCard(COLORS['neon_mint'])
        profiles_layout = QVBoxLayout(profiles_card)
        profiles_layout.setContentsMargins(2, 2, 2, 2)

        # Header
        header = QWidget()
        header.setFixedHeight(44)
        header.setStyleSheet(f"background: {COLORS['bg_darker']}; border-radius: 14px 14px 0 0;")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(16, 0, 16, 0)
        header_layout.setSpacing(12)

        # Select All
        select_widget = QWidget()
        select_layout = QHBoxLayout(select_widget)
        select_layout.setContentsMargins(0, 0, 0, 0)
        select_layout.setSpacing(8)

        self.select_all_cb = CyberCheckBox()
        self.select_all_cb.stateChanged.connect(self._toggle_select_all)
        select_layout.addWidget(self.select_all_cb)

        select_label = QLabel("Chọn tất cả")
        select_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px;")
        select_label.setCursor(Qt.PointingHandCursor)
        select_label.mousePressEvent = lambda e: self.select_all_cb.setChecked(not self.select_all_cb.isChecked())
        select_layout.addWidget(select_label)

        header_layout.addWidget(select_widget)

        sep = QFrame()
        sep.setFixedWidth(2)
        sep.setFixedHeight(24)
        sep.setStyleSheet(f"background: {COLORS['border']};")
        header_layout.addWidget(sep)

        header_title = QLabel("📁 PROFILES")
        header_title.setStyleSheet(f"color: {COLORS['neon_mint']}; font-size: 12px; font-weight: bold; letter-spacing: 2px;")
        header_layout.addWidget(header_title)

        self.count_label = QLabel("[0]")
        self.count_label.setStyleSheet(f"color: {COLORS['text_muted']}; font-size: 11px;")
        header_layout.addWidget(self.count_label)

        header_layout.addStretch()

        # Filter buttons
        self.btn_filter_all = CyberButton("Tất cả", "ghost")
        self.btn_filter_all.setFixedWidth(70)
        self.btn_filter_all.clicked.connect(lambda: self._apply_filter("all"))
        header_layout.addWidget(self.btn_filter_all)

        self.btn_filter_no_fb = CyberButton("Chưa FB", "ghost")
        self.btn_filter_no_fb.setFixedWidth(80)
        self.btn_filter_no_fb.clicked.connect(lambda: self._apply_filter("no_fb"))
        header_layout.addWidget(self.btn_filter_no_fb)

        self.btn_filter_has_fb = CyberButton("Có FB", "ghost")
        self.btn_filter_has_fb.setFixedWidth(70)
        self.btn_filter_has_fb.clicked.connect(lambda: self._apply_filter("has_fb"))
        header_layout.addWidget(self.btn_filter_has_fb)

        self.selected_label = QLabel("")
        self.selected_label.setStyleSheet(f"color: {COLORS['neon_cyan']}; font-size: 11px;")
        header_layout.addWidget(self.selected_label)

        profiles_layout.addWidget(header)

        # Table (lớn hơn)
        self.table = CyberTable(["✓", "TÊN PROFILE", "TRẠNG THÁI", "FB STATUS"])
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 250)
        self.table.setColumnWidth(2, 120)
        self.table.setColumnWidth(3, 150)

        profiles_layout.addWidget(self.table)
        content.addWidget(profiles_card, 2)  # stretch factor 2

        # RIGHT - Settings & Actions (bé hơn)
        settings_card = CyberCard(COLORS['neon_purple'])
        settings_card.setFixedWidth(300)
        settings_layout = QVBoxLayout(settings_card)
        settings_layout.setContentsMargins(12, 12, 12, 12)
        settings_layout.setSpacing(10)

        # Title
        settings_title = QLabel("⚙️ CÀI ĐẶT")
        settings_title.setStyleSheet(f"color: {COLORS['neon_purple']}; font-size: 13px; font-weight: bold; letter-spacing: 1px;")
        settings_layout.addWidget(settings_title)

        # Threads
        threads_row = QHBoxLayout()
        threads_label = QLabel("Số luồng:")
        threads_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px;")
        threads_label.setFixedWidth(70)
        threads_row.addWidget(threads_label)

        self.threads_spin = QSpinBox()
        self.threads_spin.setRange(1, 10)
        self.threads_spin.setValue(3)
        self.threads_spin.setFixedWidth(60)
        self.threads_spin.setStyleSheet(f"""
            QSpinBox {{
                background: {COLORS['bg_card']};
                color: {COLORS['text_primary']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                padding: 4px 8px;
                font-size: 12px;
            }}
        """)
        threads_row.addWidget(self.threads_spin)

        delay_label = QLabel("Delay:")
        delay_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px;")
        threads_row.addWidget(delay_label)

        self.delay_spin = QSpinBox()
        self.delay_spin.setRange(1, 60)
        self.delay_spin.setValue(5)
        self.delay_spin.setFixedWidth(60)
        self.delay_spin.setStyleSheet(self.threads_spin.styleSheet())
        threads_row.addWidget(self.delay_spin)
        threads_row.addStretch()

        settings_layout.addLayout(threads_row)

        # Dest folder
        dest_label = QLabel("Thư mục đích (LIVE):")
        dest_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px;")
        settings_layout.addWidget(dest_label)

        self.dest_combo = CyberComboBox(["-- Không chuyển --"])
        settings_layout.addWidget(self.dest_combo)

        # Options
        options_frame = QFrame()
        options_frame.setStyleSheet(f"background: {COLORS['bg_card']}; border-radius: 8px;")
        options_layout = QVBoxLayout(options_frame)
        options_layout.setContentsMargins(10, 10, 10, 10)
        options_layout.setSpacing(6)

        # Delete bad
        delete_row = QHBoxLayout()
        self.delete_cb = CyberCheckBox()
        delete_row.addWidget(self.delete_cb)
        delete_label = QLabel("Xóa profile nếu nick DIE")
        delete_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px;")
        delete_row.addWidget(delete_label)
        delete_row.addStretch()
        options_layout.addLayout(delete_row)

        # Save XLSX
        save_row = QHBoxLayout()
        self.save_cb = CyberCheckBox()
        self.save_cb.setChecked(True)
        save_row.addWidget(self.save_cb)
        save_label = QLabel("Lưu trạng thái vào XLSX")
        save_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px;")
        save_row.addWidget(save_label)
        save_row.addStretch()
        options_layout.addLayout(save_row)

        settings_layout.addWidget(options_frame)

        # Progress
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(16)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                background: {COLORS['bg_darker']};
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, x2:1, stop:0 {COLORS['neon_mint']}, stop:1 {COLORS['neon_cyan']});
                border-radius: 7px;
            }}
        """)
        self.progress_bar.setValue(0)
        settings_layout.addWidget(self.progress_bar)

        self.progress_label = QLabel("")
        self.progress_label.setStyleSheet(f"color: {COLORS['neon_mint']}; font-size: 11px;")
        self.progress_label.setAlignment(Qt.AlignCenter)
        settings_layout.addWidget(self.progress_label)

        settings_layout.addStretch()

        # Actions
        actions_title = QLabel("🚀 HÀNH ĐỘNG")
        actions_title.setStyleSheet(f"color: {COLORS['neon_pink']}; font-size: 12px; font-weight: bold; letter-spacing: 1px;")
        settings_layout.addWidget(actions_title)

        btn_check = CyberButton("KIỂM TRA FB", "cyan", "🔍")
        btn_check.clicked.connect(self._check_fb_status)
        settings_layout.addWidget(btn_check)

        btn_start = CyberButton("BẮT ĐẦU ĐĂNG NHẬP", "success", "▶️")
        btn_start.clicked.connect(self._start_login)
        settings_layout.addWidget(btn_start)

        btn_stop = CyberButton("DỪNG LẠI", "danger", "⏹️")
        btn_stop.clicked.connect(self._stop_login)
        settings_layout.addWidget(btn_stop)

        content.addWidget(settings_card)
        layout.addLayout(content, 1)

    def _load_folders(self):
        """Tải danh sách thư mục từ Hidemium"""
        self.log("Đang tải thư mục...", "info")

        def fetch():
            try:
                return api.get_folders(limit=100)
            except Exception as e:
                self.log(f"Lỗi kết nối Hidemium: {e}", "error")
                return []

        def on_complete(folders):
            self.folders = folders or []

            self.folder_combo.clear()
            self.folder_combo.addItem("📁 Chọn thư mục")
            for f in self.folders:
                name = f.get('name', 'Không rõ')
                self.folder_combo.addItem(f"📁 {name}")

            self.dest_combo.clear()
            self.dest_combo.addItem("-- Không chuyển --")
            for f in self.folders:
                name = f.get('name', 'Không rõ')
                self.dest_combo.addItem(f"📁 {name}")

            self.log(f"Đã tải {len(self.folders)} thư mục", "success")

        def run():
            result = fetch()
            QTimer.singleShot(0, lambda: on_complete(result))

        threading.Thread(target=run, daemon=True).start()

    def _on_folder_change(self, index):
        if index > 0:
            self._load_profiles()

    def _load_profiles(self):
        """Tải profiles từ thư mục"""
        folder_idx = self.folder_combo.currentIndex()
        if folder_idx <= 0:
            return

        folder = self.folders[folder_idx - 1]
        folder_id = folder.get('id')

        self.log(f"Đang tải profiles từ {folder.get('name')}...", "info")

        def fetch():
            try:
                return api.get_profiles(folder_id=[folder_id], limit=500)
            except:
                return []

        def on_complete(profiles):
            self.profiles = profiles or []
            self._update_table()
            self._update_stats()
            self.log(f"Đã tải {len(self.profiles)} profiles", "success")

        def run():
            result = fetch()
            QTimer.singleShot(0, lambda: on_complete(result))

        threading.Thread(target=run, daemon=True).start()

    def _update_table(self, filter_type: str = "all"):
        """Cập nhật bảng với profiles"""
        self.profile_checkboxes.clear()

        profiles_to_show = self.profiles
        if filter_type == "no_fb":
            profiles_to_show = [p for p in self.profiles
                               if not self.profile_status.get(p.get('uuid'), {}).get('has_fb', False)]
        elif filter_type == "has_fb":
            profiles_to_show = [p for p in self.profiles
                               if self.profile_status.get(p.get('uuid'), {}).get('has_fb', False)]

        self.table.setRowCount(len(profiles_to_show))

        for row, profile in enumerate(profiles_to_show):
            uuid = profile.get('uuid', '')
            name = profile.get('name', 'Không rõ')
            status = profile.get('status', 'stopped')
            fb_info = self.profile_status.get(uuid, {})
            has_fb = fb_info.get('has_fb')

            # Checkbox
            cb_widget = QWidget()
            cb_widget.setStyleSheet("background: transparent;")
            cb_layout = QHBoxLayout(cb_widget)
            cb_layout.setContentsMargins(0, 0, 0, 0)
            cb_layout.setAlignment(Qt.AlignCenter)
            checkbox = CyberCheckBox()
            checkbox.stateChanged.connect(self._update_selection_count)
            cb_layout.addWidget(checkbox)
            self.table.setCellWidget(row, 0, cb_widget)
            self.profile_checkboxes[uuid] = checkbox

            # Name
            name_item = QTableWidgetItem(name[:30] + "..." if len(name) > 30 else name)
            self.table.setItem(row, 1, name_item)

            # Status
            status_text = "🟢 Đang chạy" if status == "running" else "⚫ Đã dừng"
            status_item = QTableWidgetItem(status_text)
            self.table.setItem(row, 2, status_item)

            # FB Status
            if has_fb is True:
                fb_text = "✅ Có FB"
                fb_color = COLORS['neon_mint']
            elif has_fb is False:
                fb_text = "❌ Chưa có FB"
                fb_color = COLORS['neon_coral']
            else:
                fb_text = "❓ Chưa kiểm tra"
                fb_color = COLORS['text_muted']

            fb_item = QTableWidgetItem(fb_text)
            fb_item.setForeground(QColor(fb_color))
            self.table.setItem(row, 3, fb_item)

        self.count_label.setText(f"[{len(profiles_to_show)} profiles]")

    def _apply_filter(self, filter_type: str):
        """Áp dụng bộ lọc cho bảng"""
        self._update_table(filter_type)

    def _toggle_select_all(self, state):
        checked = state == Qt.Checked
        for uuid, cb in self.profile_checkboxes.items():
            cb.blockSignals(True)
            cb.setChecked(checked)
            cb.blockSignals(False)
        self._update_selection_count()

    def _update_selection_count(self):
        count = sum(1 for cb in self.profile_checkboxes.values() if cb.isChecked())
        self.selected_label.setText(f"✓ {count} đã chọn" if count > 0 else "")
        self.stat_selected.set_value(str(count))

    def _update_stats(self):
        self.stat_profiles.set_value(str(len(self.profiles)))
        selected = sum(1 for cb in self.profile_checkboxes.values() if cb.isChecked())
        self.stat_selected.set_value(str(selected))
        unused = len([a for a in self.accounts if not a.get('status')])
        self.stat_accounts.set_value(str(unused))

        # Count LIVE and DIE
        live_count = sum(1 for s in self.profile_status.values() if s.get('result') == 'LIVE')
        die_count = sum(1 for s in self.profile_status.values() if s.get('result') in ['DIE', 'WRONG_PASS', 'LOCKED'])
        self.stat_live.set_value(str(live_count))
        self.stat_die.set_value(str(die_count))

    def _browse_xlsx(self):
        """Chọn file XLSX"""
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Lỗi", "Cần cài đặt openpyxl:\npip install openpyxl")
            return

        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn file XLSX", "", "Excel Files (*.xlsx)"
        )

        if path:
            self.xlsx_path = path
            self.xlsx_input.setText(os.path.basename(path))
            self._load_xlsx()

    def _load_xlsx(self):
        """Tải tài khoản từ XLSX - Format: A(Status), B(UID), C(Password), D(2FA), E(Email), F(EmailPass)"""
        try:
            self.workbook = openpyxl.load_workbook(self.xlsx_path)
            sheet = self.workbook.active

            self.accounts = []

            def clean_str(val):
                if val is None:
                    return ''
                if isinstance(val, float):
                    return str(int(val)) if val == int(val) else str(val)
                return str(val)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                if not row or len(row) < 3:
                    continue

                account = {
                    'row': row_idx,
                    'status': clean_str(row[0]),
                    'fb_id': clean_str(row[1]),
                    'password': clean_str(row[2]),
                    'totp_secret': clean_str(row[3]) if len(row) > 3 else '',
                    'email': clean_str(row[4]) if len(row) > 4 else '',
                    'email_pass': clean_str(row[5]) if len(row) > 5 else ''
                }

                if account['fb_id']:
                    self.accounts.append(account)

            unused = [a for a in self.accounts if not a['status']]
            self.stat_accounts.set_value(str(len(unused)))
            self.log(f"Đã import {len(self.accounts)} tài khoản ({len(unused)} chưa dùng)", "success")

        except Exception as e:
            self.log(f"Lỗi đọc file XLSX: {e}", "error")

    def _update_xlsx_status(self, row: int, status: str):
        """Cập nhật trạng thái vào file XLSX"""
        if not self.workbook or not self.save_cb.isChecked():
            return

        try:
            sheet = self.workbook.active
            sheet.cell(row=row, column=1, value=status)
            self.workbook.save(self.xlsx_path)
        except Exception as e:
            print(f"Lỗi lưu XLSX: {e}")

    def _get_dest_folder_uuid(self) -> Optional[str]:
        """Lấy uuid của thư mục đích"""
        dest_idx = self.dest_combo.currentIndex()
        if dest_idx <= 0:
            return None
        folder = self.folders[dest_idx - 1]
        return folder.get('uuid') or folder.get('id')

    def _check_fb_status(self):
        """Kiểm tra profiles đã có FB chưa"""
        if self._is_running:
            return

        selected = [uuid for uuid, cb in self.profile_checkboxes.items() if cb.isChecked()]
        if not selected:
            QMessageBox.warning(self, "Lỗi", "Chưa chọn profile nào!")
            return

        if not HAS_WEBSOCKET:
            QMessageBox.warning(self, "Lỗi", "Cần cài đặt websocket:\npip install websocket-client requests")
            return

        self._is_running = True
        self._stop_requested = False

        thread_count = self.threads_spin.value()
        total = len(selected)

        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(0)
        self.log(f"Bắt đầu kiểm tra FB cho {total} profiles...", "info")

        def check_worker(profile_queue: queue.Queue, progress: list):
            while not self._stop_requested:
                try:
                    uuid = profile_queue.get_nowait()
                except queue.Empty:
                    break

                slot_id = acquire_window_slot()
                try:
                    result = api.open_browser(uuid)
                    if result.get('status') == 'successfully':
                        data = result.get('data', {})
                        remote_port = data.get('remote_port')

                        # Set window bounds
                        self._set_browser_bounds(remote_port, slot_id)

                        # Check FB
                        has_fb = self._check_profile_has_fb(remote_port)
                        self.profile_status[uuid] = {'has_fb': has_fb}

                        api.close_browser(uuid)

                        name = next((p.get('name', '') for p in self.profiles if p.get('uuid') == uuid), uuid[:8])
                        status_text = "✅ Có FB" if has_fb else "❌ Chưa có FB"
                        QTimer.singleShot(0, lambda n=name, s=status_text: self.log(f"[{n}] {s}", "info"))

                except Exception as e:
                    QTimer.singleShot(0, lambda err=str(e): self.log(f"Lỗi: {err}", "error"))
                finally:
                    release_window_slot(slot_id)

                progress[0] += 1
                QTimer.singleShot(0, lambda v=progress[0]: self.progress_bar.setValue(v))
                profile_queue.task_done()

        def run_check():
            q = queue.Queue()
            for uuid in selected:
                q.put(uuid)

            progress = [0]
            threads = []
            for _ in range(min(thread_count, len(selected))):
                t = threading.Thread(target=check_worker, args=(q, progress))
                t.start()
                threads.append(t)

            for t in threads:
                t.join()

            self._is_running = False
            QTimer.singleShot(0, self._update_table)
            QTimer.singleShot(0, self._update_stats)
            QTimer.singleShot(0, lambda: self.log("Kiểm tra FB hoàn thành!", "success"))
            QTimer.singleShot(0, lambda: self.progress_label.setText("Hoàn thành!"))

        threading.Thread(target=run_check, daemon=True).start()

    def _set_browser_bounds(self, remote_port: int, slot_id: int):
        """Đặt vị trí cửa sổ trình duyệt theo slot"""
        import json as json_module

        try:
            resp = requests.get(f"http://127.0.0.1:{remote_port}/json", timeout=5)
            tabs = resp.json()

            page_ws = None
            for tab in tabs:
                if tab.get('type') == 'page':
                    page_ws = tab.get('webSocketDebuggerUrl')
                    break

            if not page_ws:
                return

            ws = websocket.create_connection(page_ws, timeout=5, suppress_origin=True)
            x, y, w, h = get_window_bounds(slot_id)

            ws.send(json_module.dumps({"id": 1, "method": "Browser.getWindowForTarget", "params": {}}))
            result = json_module.loads(ws.recv())

            if result and 'result' in result and 'windowId' in result['result']:
                window_id = result['result']['windowId']
                ws.send(json_module.dumps({
                    "id": 2,
                    "method": "Browser.setWindowBounds",
                    "params": {
                        "windowId": window_id,
                        "bounds": {"left": x, "top": y, "width": w, "height": h, "windowState": "normal"}
                    }
                }))
                ws.recv()

            ws.close()
        except Exception as e:
            print(f"Lỗi đặt kích thước: {e}")

    def _check_profile_has_fb(self, remote_port: int) -> bool:
        """Kiểm tra trình duyệt đã đăng nhập FB chưa qua CDP"""
        import json as json_module

        try:
            time.sleep(1.5)
            resp = requests.get(f"http://127.0.0.1:{remote_port}/json", timeout=5)
            tabs = resp.json()

            page_ws = None
            for tab in tabs:
                if tab.get('type') == 'page':
                    page_ws = tab.get('webSocketDebuggerUrl')
                    break

            if not page_ws:
                return False

            ws = websocket.create_connection(page_ws, timeout=15, suppress_origin=True)

            # Navigate to Facebook
            ws.send(json_module.dumps({
                "id": 1, "method": "Page.navigate",
                "params": {"url": "https://www.facebook.com"}
            }))
            ws.recv()
            time.sleep(3)

            # Check if logged in
            js = '''
                (function() {
                    let loginForm = document.querySelector('input[name="email"], input[name="pass"]');
                    if (loginForm) return false;

                    let profileMenu = document.querySelector('[aria-label*="Account"], [aria-label*="Tài khoản"]');
                    let messenger = document.querySelector('[aria-label*="Messenger"]');
                    let notifications = document.querySelector('[aria-label*="Notifications"], [aria-label*="Thông báo"]');

                    return !!(profileMenu || messenger || notifications);
                })()
            '''

            ws.send(json_module.dumps({
                "id": 2, "method": "Runtime.evaluate",
                "params": {"expression": js, "returnByValue": True}
            }))
            result = json_module.loads(ws.recv())
            ws.close()

            return result.get('result', {}).get('result', {}).get('value', False) is True

        except Exception as e:
            print(f"Lỗi kiểm tra FB: {e}")
            return False

    def _start_login(self):
        """Bắt đầu đăng nhập - FULL LOGIC"""
        if self._is_running:
            return

        selected = [uuid for uuid, cb in self.profile_checkboxes.items() if cb.isChecked()]
        unused = [a for a in self.accounts if not a.get('status')]

        if not selected:
            QMessageBox.warning(self, "Lỗi", "Chưa chọn profile nào!")
            return

        if not unused:
            QMessageBox.warning(self, "Lỗi", "Không có tài khoản chưa dùng!")
            return

        if not HAS_WEBSOCKET:
            QMessageBox.warning(self, "Lỗi", "Cần cài đặt websocket:\npip install websocket-client requests")
            return

        self._is_running = True
        self._stop_requested = False

        thread_count = self.threads_spin.value()
        delay = self.delay_spin.value()
        total = min(len(selected), len(unused))

        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(0)
        self.log(f"Bắt đầu đăng nhập {total} profiles...", "info")

        def login_worker(profile_queue: queue.Queue, account_queue: queue.Queue, progress: list):
            import json as json_module

            while not self._stop_requested:
                try:
                    uuid = profile_queue.get_nowait()
                except queue.Empty:
                    break

                profile = next((p for p in self.profiles if p.get('uuid') == uuid), None)
                if not profile:
                    profile_queue.task_done()
                    continue

                profile_name = profile.get('name', uuid[:8])

                try:
                    account = account_queue.get_nowait()
                except queue.Empty:
                    QTimer.singleShot(0, lambda pn=profile_name: self.log(f"[{pn}] Hết tài khoản", "warning"))
                    profile_queue.task_done()
                    break

                QTimer.singleShot(0, lambda pn=profile_name, fb=account['fb_id'][:12]:
                    self.log(f"[{pn}] Đăng nhập với {fb}...", "info"))

                slot_id = acquire_window_slot()

                try:
                    # Open browser
                    result = api.open_browser(uuid)
                    if result.get('status') != 'successfully':
                        account['status'] = 'ERROR'
                        self._update_xlsx_status(account['row'], 'ERROR')
                        release_window_slot(slot_id)
                        profile_queue.task_done()
                        continue

                    data = result.get('data', {})
                    remote_port = data.get('remote_port')

                    if not remote_port:
                        account['status'] = 'NO_PORT'
                        self._update_xlsx_status(account['row'], 'NO_PORT')
                        release_window_slot(slot_id)
                        profile_queue.task_done()
                        continue

                    # Set window bounds
                    self._set_browser_bounds(remote_port, slot_id)

                    # Get page WebSocket
                    time.sleep(1)
                    resp = requests.get(f"http://127.0.0.1:{remote_port}/json", timeout=5)
                    tabs = resp.json()

                    page_ws = None
                    for tab in tabs:
                        if tab.get('type') == 'page':
                            page_ws = tab.get('webSocketDebuggerUrl')
                            break

                    if not page_ws:
                        account['status'] = 'NO_PAGE'
                        self._update_xlsx_status(account['row'], 'NO_PAGE')
                        api.close_browser(uuid)
                        release_window_slot(slot_id)
                        profile_queue.task_done()
                        continue

                    # Connect WebSocket
                    ws = websocket.create_connection(page_ws, timeout=15, suppress_origin=True)
                    msg_id = [0]

                    def send_cmd(method, params=None):
                        msg_id[0] += 1
                        msg = {"id": msg_id[0], "method": method}
                        if params:
                            msg["params"] = params
                        ws.send(json_module.dumps(msg))
                        return json_module.loads(ws.recv())

                    def evaluate(expression):
                        result = send_cmd("Runtime.evaluate", {
                            "expression": expression, "returnByValue": True
                        })
                        return result.get('result', {}).get('result', {}).get('value')

                    # Clear cookies
                    for domain in [".facebook.com", "facebook.com", "www.facebook.com"]:
                        send_cmd("Network.deleteCookies", {"domain": domain})

                    # Navigate to login
                    send_cmd("Page.navigate", {"url": "https://www.facebook.com/login"})
                    time.sleep(3)

                    # Type credentials
                    def type_text(text, selector):
                        evaluate(f'''
                            (function() {{
                                let el = document.querySelector('{selector}');
                                if (el) {{ el.focus(); el.value = ''; }}
                            }})()
                        ''')
                        time.sleep(0.3)
                        for char in text:
                            send_cmd("Input.insertText", {"text": char})
                            time.sleep(random.uniform(0.1, 0.25))

                    type_text(account['fb_id'], "#email")
                    time.sleep(0.5)
                    type_text(account['password'], "#pass")
                    time.sleep(0.5)

                    # Click login
                    evaluate('''
                        (function() {
                            let btn = document.querySelector('#loginbutton, button[name="login"], button[type="submit"]');
                            if (btn) btn.click();
                        })()
                    ''')

                    time.sleep(4)

                    # Check result
                    js_check = '''
                        (function() {
                            let url = window.location.href;
                            let pageText = (document.body.innerText || '').toLowerCase();

                            if (url.includes('checkpoint') || url.includes('locked')) return 'LOCKED';
                            if (url.includes('two_step') || document.querySelector('input[name="approvals_code"]')) return '2FA';

                            let wrongPhrases = ['password that you', 'incorrect password', 'sai mật khẩu'];
                            for (let p of wrongPhrases) if (pageText.includes(p)) return 'WRONG_PASS';

                            let diePhrases = ['vô hiệu hóa', 'disabled', 'suspended'];
                            for (let p of diePhrases) if (pageText.includes(p)) return 'DIE';

                            let isLoggedIn = document.querySelector('[aria-label*="Account"], [aria-label*="Messenger"]');
                            if (isLoggedIn || url === 'https://www.facebook.com/') return 'LIVE';

                            return 'UNKNOWN';
                        })()
                    '''

                    status = evaluate(js_check) or 'UNKNOWN'

                    # Handle 2FA
                    if status == '2FA' and account.get('totp_secret'):
                        totp_code = self._generate_totp(account['totp_secret'])
                        QTimer.singleShot(0, lambda c=totp_code: self.log(f"  Mã 2FA: {c}", "info"))

                        evaluate('''
                            (function() {
                                let input = document.querySelector('input[name="approvals_code"], input[type="text"]');
                                if (input) input.focus();
                            })()
                        ''')
                        time.sleep(0.5)

                        for digit in totp_code:
                            send_cmd("Input.insertText", {"text": digit})
                            time.sleep(0.2)

                        evaluate('''
                            (function() {
                                let btns = document.querySelectorAll('div[role="button"], button');
                                for (let btn of btns) {
                                    let text = (btn.innerText || '').toLowerCase();
                                    if (text.includes('tiếp tục') || text.includes('continue')) {
                                        btn.click(); break;
                                    }
                                }
                            })()
                        ''')

                        time.sleep(3)
                        status = evaluate(js_check) or 'UNKNOWN'

                    ws.close()

                    # Update status
                    account['status'] = status
                    self._update_xlsx_status(account['row'], status)
                    self.profile_status[uuid] = {'has_fb': status == 'LIVE', 'result': status}

                    # Log result
                    if status == 'LIVE':
                        QTimer.singleShot(0, lambda pn=profile_name: self.log(f"[{pn}] ✅ LIVE", "success"))

                        # Move to dest folder
                        dest_folder = self._get_dest_folder_uuid()
                        if dest_folder:
                            try:
                                api.add_profiles_to_folder(dest_folder, [uuid])
                                QTimer.singleShot(0, lambda pn=profile_name: self.log(f"[{pn}] Đã chuyển đến thư mục đích", "info"))
                            except Exception as e:
                                pass

                        time.sleep(random.uniform(2, 5))
                        api.close_browser(uuid)
                    else:
                        QTimer.singleShot(0, lambda pn=profile_name, s=status: self.log(f"[{pn}] ❌ {s}", "error"))
                        api.close_browser(uuid)

                        # Delete profile if option checked
                        if self.delete_cb.isChecked() and status in ['DIE', 'WRONG_PASS', 'LOCKED']:
                            try:
                                api.delete_profiles([uuid], is_local=True)
                                QTimer.singleShot(0, lambda pn=profile_name: self.log(f"[{pn}] Đã xóa profile", "warning"))
                            except:
                                pass

                except Exception as e:
                    QTimer.singleShot(0, lambda pn=profile_name, err=str(e): self.log(f"[{pn}] Lỗi: {err}", "error"))
                    account['status'] = 'ERROR'
                    self._update_xlsx_status(account['row'], 'ERROR')
                    try:
                        api.close_browser(uuid)
                    except:
                        pass

                finally:
                    release_window_slot(slot_id)

                progress[0] += 1
                QTimer.singleShot(0, lambda v=progress[0]: self.progress_bar.setValue(v))
                QTimer.singleShot(0, lambda v=progress[0], t=total: self.progress_label.setText(f"{v}/{t}"))

                time.sleep(delay)
                profile_queue.task_done()

        def run_login():
            profile_q = queue.Queue()
            account_q = queue.Queue()

            for uuid in selected:
                profile_q.put(uuid)
            for acc in unused:
                account_q.put(acc)

            progress = [0]
            threads = []
            for _ in range(min(thread_count, len(selected))):
                t = threading.Thread(target=login_worker, args=(profile_q, account_q, progress))
                t.start()
                threads.append(t)

            for t in threads:
                t.join()

            self._is_running = False
            QTimer.singleShot(0, self._update_table)
            QTimer.singleShot(0, self._update_stats)
            QTimer.singleShot(0, lambda: self.log("Đăng nhập hoàn thành!", "success"))
            QTimer.singleShot(0, lambda: self.progress_label.setText("Hoàn thành!"))

        threading.Thread(target=run_login, daemon=True).start()

    def _generate_totp(self, secret: str) -> str:
        """Tạo mã TOTP 6 chữ số"""
        import hmac
        import hashlib
        import struct

        secret = secret.replace(' ', '').upper()
        base32_chars = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ234567'
        bits = ''
        for c in secret:
            if c in base32_chars:
                bits += bin(base32_chars.index(c))[2:].zfill(5)
        key = bytes(int(bits[i:i+8], 2) for i in range(0, len(bits) - len(bits) % 8, 8))

        counter = int(time.time()) // 30
        counter_bytes = struct.pack('>Q', counter)

        hmac_hash = hmac.new(key, counter_bytes, hashlib.sha1).digest()
        offset = hmac_hash[-1] & 0x0F
        code = struct.unpack('>I', hmac_hash[offset:offset+4])[0] & 0x7FFFFFFF
        return str(code % 1000000).zfill(6)

    def _stop_login(self):
        """Dừng đăng nhập"""
        self._stop_requested = True
        self.log("Đang dừng...", "warning")
        self.progress_label.setText("Đang dừng...")
